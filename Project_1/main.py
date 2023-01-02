from bs4 import BeautifulSoup
import requests
import openpyxl

excel=openpyxl.Workbook()
# Excel got created
print(excel.sheetnames)
# Printing the names of the sheet. 
sheet=excel.active
sheet.title='Top Rated Movies'
print(excel.sheetnames)
# Name of the excel file got changed. 

# Name of the columns of the excel file. 
sheet.append(['Movie Rank','Movie Name','Movie Year','IMDB Rating'])

# source=requests.get("https://www.imdb.com/chart/top/")

# This get method will give a response object which will be stored inside the source variable. 

# Response object will be having the html source code of the gived web page. 

# HTML Code can be taken from both right click and view source page as well as rt click and inspect. 

# Now it is a good practice to use requests parameter in try exception block to avoid any issue while usinf a wrong link. 

try:
    source=requests.get("https://www.imdb.com/chart/top/")
    source.raise_for_status()
    soup=BeautifulSoup(source.text,'html.parser') #Taking html text of the response object. html parser is the default parser comes with python installation but we can use someother parser too.
    # This will return a BeautifulSoup object which will be retained in soup 
    # print(soup)

    # Now as we know that by clicking right click on the movie name an then inspect we can see that we are being directed to the place where the name was called. 
    # Now the place where i is called we can check the parent tag and then we can find with soup from parent tag onwards.
    # It is a good practice so that we dont get lost in the page.
    # Here we are having a table where 1 tr tag is for 1 movie having multiple td tag with names rating etc.
    # The name where we right click was in 1 td tag so when we go upwards we close tr followed by a tag called as tbody which is the mail parent tag for all the movies.

    movies1=soup.find('tbody',class_='lister-list')
    print(type(movies1))

    # Here we have only 1 tbody tag in the page hence its ok or else find function will only find the first one. 
    # Now we need to find allt he tr tag and then the td tag but tr tags are in huge number hence we need ti use a function find all
    # find all function will return a generator object which can be easily iterated and we can take the value of td tag out .

    # movies2=movies.find_all('tr')
    # print(len(movies2))

    # This is 1 way to do this there are other ways that is: 
    movies=soup.find('tbody',class_='lister-list').find_all('tr')
    # As tr have no class hence all tr needed in tbody tag. 
    print(len(movies))
    # print(movies)
    print(type(movies)) 
    # This is also working. 
    # Initially movies was an object storing a tag in text format but now it is a resultset object which can be interated. 
    # Iterating the result set object. .
    for movie in movies:
        # print(movie)
        # break
        # Here we are just checking what is comming out in movie object it should give only 1st tr tag. It is correct now we need to go to 2nd tr for movie name. 
        # name=movie.find('td',class_='titleColumn')
        # print(name)
        # break
        # Here we are printing the complete td tag for the name 
        # Now we need to just have the value or the text so we can just add the tag name in find and as there is only 1 anchor tag hence no need to afdd class name. 
        # name=movie.find('td',class_='titleColumn').a
        # print(name)
        # break
        # Now here we can see that the complete detail of anchor tag is  filled now we just need the text.
        # name=movie.find('td','titleColumn').a.text
        # print(name)
        # break
        # As we can see now the daat is correct. 
        name=movie.find('td',class_='titleColumn').a.text
        # print(name)

        # Now to get rank we need to check the tag so again inspect this and we can see that td tag is having value as 1. 
        # rank1=movie.find('td',class_='titleColumn').text
        # print(rank1)

        # Here we can see that all the data in the td tag for class titleColumn is being printed. 

        # rank2=movie.find('td',class_='titleColumn').get_text(strip=True)
        # print(rank2)

        #Here we are getting the value and all the sapces are removed this is caused due to get_text function as we can send some parameters.
        # strip=True removes all extra spaces. 

        # Now as we can see that after 1 we have a . symbol hence it can be used to split the string. 
        # Here we use a function split('parameter by which to split'). THIS WILL RETURN A LIST SPLITTING THE ELEMENT INTO 2 PARTS. 

        # rank3=movie.find('td',class_='titleColumn').get_text(strip=True).split('.')
        # print(rank3)

        # Now to get 1 we just need to get the 0th elkement or the 1st element of the list. so just pass[0]

        rank=movie.find('td',class_='titleColumn').get_text(strip=True).split('.')[0]
        # print(rank)

        # Similarly for the year we need to go to td then go to span and get the text. 

        # year1=movie.find('td',class_='titleColumn').find('span',class_='secondaryInfo').text.split('(')[1].split(')')[0]
        # print(year1)

        # This way is working fine but there is a better way to do that. 
        year=movie.find('td',class_='titleColumn').find('span',class_='secondaryInfo').text.strip('()')
        # print(year)

        # Now we need imdb raitng: rt click on rating and inspect inside a different td and inside strong tag
        # Copy clas name: rt click and attribute and copy attribute. 
        rating=movie.find('td',class_='ratingColumn imdbRating').strong.text
        # print(rating)
        
        # Printing for all 250 movies now.

        print(rank,name,year,rating)
        sheet.append([rank,name,year,rating])
        # This is used to get the values in excel file. Passed as an list object

# The function raise_for_status() is used so that the error thrown by the response or requests is being printed. 

except Exception as e:
    print(e)

# Saving the excel is needed: The name is the name given to the file
excel.save('Top IMDB Rated Movies.xlsx') 

# EXCEL LOADING:

# For this first we need to import a package: openpyxl Before that pip install openpyxl needed
# Then we need to create anew excel 
# Then we need to check how many sheet that excel have : mostly 1 to 3 
# Then we need to gert tot he active excel sheet: excel.active 
# Then we can provide title or file name to excel using .title command
# Then we can also provisde the excel columns by using append and then passing a list 
# Then in try block just after printing each and every movie we can just save them in excel uisng same .append function and pass the values we are printing.
# Then finally we save the excel file and then in the project we will get a new excel file getting generated. 